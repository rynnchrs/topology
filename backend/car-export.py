import os
import sys

import django

sys.path.append(os.path.join(os.path.dirname(__file__), 'backend'))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'backend.settings')
django.setup()

from django.conf import settings
from django.core.exceptions import ObjectDoesNotExist
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from car.models import TPL, Car, Contract, Insurance

datas = Car.objects.all()
workbook = Workbook()

# Get active worksheet/tab
worksheet = workbook.active
worksheet.title = 'Car-Inventory.xlsx'

cell_1 = ["B1","F1","L1","S1","AH1","AO1","AS1","AX1","BM1","BU1","CC1","CL1"]
cell_2 = ["E1","K1","R1","AG1","AN1","AR1","AW1","BL1","BT1","CB1","CK1","CT1"]
header = ["IDENTIFICATION","VEHICLE INFOMATION","SUPPLIERS","ENGINE AND BODY INFORMATION","LTO",
        "LOCATION","DELIVERY INFO","RECEIVED ITEMS","BIDDING & CONTRACT","2019 THIRD-PARTY LIABILITY",
        "2019 COMPREHENSIVE INSURANCE","2020 COMPREHENSIVE INSURANCE"]
color = ["00FFFF00","003366FF","00FFCC99","00C0C0C0","0000FF00","00FF0000","00800080","00808000",
        "00008000","000066CC","00FF6600","00FF6600"]

for i in range(12):
    worksheet.merge_cells(cell_1[i] +":"+cell_2[i])

    cell = worksheet[cell_1[i]] 
    cell.value = header[i]
    
    cell.fill = PatternFill("solid", fgColor=color[i])
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(bold=True)
    thin = Side(border_style="thin", color="000000")
    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
# Define the titles for columns
columns = [
    "ID",
    "Vin No.",
    "Body No.",
    "CS No.",
    "Plate No.",
    "Brand",
    "Year",
    "Make",
    "Series",
    "Body Type",
    "Color",
    "Dealer",
    "Dealer Phone",
    "Dealer Email",
    "PO #",
    "PO Date",
    "Body Builder",
    "Fabricator",
    "Sale Price",
    "Vat Price",
    "Engine #",
    "Battery #",
    "Fuel Type",
    "Transmission",
    "Denomination",
    "Piston",
    "Cylinder",
    "Pocuring Entity",
    "Capacity",
    "Gross Wt.",
    "Net Wt.",
    "Shippping Wt.",
    "Net Capacity",
    "LTO CR",
    "CR Date",
    "O.R. No.",
    "O.R. Date",
    "TopLoadReg",
    "Field Office",
    "ORCR Copy",
    "Permanent",
    "Current",
    "With VTF?",
    "Permanent?",
    "Delivery Location",
    "Delivery Date",
    "SI No.",
    "DR #",
    "DR Codes",
    "Plate No. Delivery",
    "Decals",
    "Modified",
    "EWD",
    "Tools",
    "User's Manual",
    "Warranty Book",
    "Unit Key",
    "Body Key",
    "Cigarette Plug",
    "Key Chain",
    "Fan",
    "Jack",
    "Tire Wrench",
    "Fire Extinguisher",

    "Client Name",
    "Contract No.",
    "Start Date",
    "End Date",
    "Bid No.",
    "Bid Name",
    "Bid Date",
    "Bid Cost",

    "Insurance Company",
    "Telephone",
    "Email",
    "Policy No.",
    "Date Issued",
    "From",
    "To",
    "Cost",

    "Insurance Company",
    "Telephone",
    "Email",
    "Policy No.",
    "Reference No.",
    "Date Issued",
    "From",
    "To",
    "Cost",

    "Insurance Company",
    "Telephone",
    "Email",
    "Policy No.",
    "Reference No.",
    "Date Issued",
    "From",
    "To",
    "Cost",
    
    "Remarks",
    "Operational",
    "Status",
    "Date Updated",
    "Date Created"
]
row_num = 2

# Assign the titles for each cell of the header
for col_num, column_title in enumerate(columns, 1):
    cell = worksheet.cell(row=row_num, column=col_num)
    cell.value = column_title
    cell.fill = PatternFill("solid", fgColor=color[i])
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(bold=True)
    thin = Side(border_style="thin", color="000000")
    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    if col_num >= 2:
        cell.fill = PatternFill("solid", fgColor=color[0])
    if col_num >= 6:
        cell.fill = PatternFill("solid", fgColor=color[1])
    if col_num >= 12:
        cell.fill = PatternFill("solid", fgColor=color[2])
    if col_num >= 19:
        cell.fill = PatternFill("solid", fgColor=color[3])
    if col_num >= 34:
        cell.fill = PatternFill("solid", fgColor=color[4])
    if col_num >= 41:
        cell.fill = PatternFill("solid", fgColor=color[5])
    if col_num >= 45:
        cell.fill = PatternFill("solid", fgColor=color[6])
    if col_num >= 50:
        cell.fill = PatternFill("solid", fgColor=color[7])
    if col_num >= 65:
        cell.fill = PatternFill("solid", fgColor=color[8])
    if col_num >= 73:
        cell.fill = PatternFill("solid", fgColor=color[9])
    if col_num >= 81:
        cell.fill = PatternFill("solid", fgColor=color[10])
    if col_num >= 90:
        cell.fill = PatternFill("solid", fgColor=color[11])
    if col_num >= 99:
        cell.fill = PatternFill("solid", fgColor="00FFFFFF")

# Iterate through all movies
for data in datas:
    row_num += 1
    try:
        contract = Contract.objects.get(car=data.car_id)
    except ObjectDoesNotExist:
        contract = None
        contract_client_name = ""
        contract_contract_no = ""
        contract_start_date = ""
        contract_end_date = ""
        contract_bid_no = ""
        contract_bid_name = ""
        contract_bid_date = ""
        contract_cost = ""
    if contract is not None:
        contract_client_name = contract.client_name
        contract_contract_no = contract.contract_no
        contract_start_date = contract.start_date
        contract_end_date = contract.end_date
        contract_bid_no = contract.bid_no
        contract_bid_name = contract.bid_name
        contract_bid_date = contract.bid_date
        contract_cost = contract.cost
    try:
        tpl = TPL.objects.get(car=data.car_id)
    except ObjectDoesNotExist:
        tpl = None
        tpl_insurance_name = ""
        tpl_telephone = ""
        tpl_email = ""
        tpl_po_no = ""
        tpl_date_issued = ""
        tpl_start_date = ""
        tpl_end_date = ""
        tpl_cost = ""
    if tpl is not None:
        tpl_insurance_name = tpl.insurance_name
        tpl_telephone = tpl.telephone
        tpl_email = tpl.email
        tpl_po_no = tpl.po_no
        tpl_date_issued = tpl.date_issued
        tpl_start_date = tpl.start_date
        tpl_end_date = tpl.end_date
        tpl_cost = tpl.cost

    try:
        ins19 = Insurance.objects.get(car=data.car_id, insurance_no = 1)
    except ObjectDoesNotExist:
        ins19 = None
        ins19_company = ""
        ins19_telephone = ""
        ins19_email = ""
        ins19_po_no = ""
        ins19_reference_no = ""
        ins19_date_issued = ""
        ins19_start_date = ""
        ins19_end_date = ""
        ins19_cost = ""
    if ins19 is not None:
        ins19_company = ins19.company
        ins19_telephone = ins19.telephone
        ins19_email = ins19.email
        ins19_po_no = ins19.po_no
        ins19_reference_no = ins19.reference_no
        ins19_date_issued = ins19.date_issued
        ins19_start_date = ins19.start_date
        ins19_end_date = ins19.end_date
        ins19_cost = ins19.cost
    try:
        ins20 = Insurance.objects.get(car=data.car_id, insurance_no = 2)
    except ObjectDoesNotExist:
        ins20 = None
        ins20_company = ""
        ins20_telephone = ""
        ins20_email = ""
        ins20_po_no = ""
        ins20_reference_no = ""
        ins20_date_issued = ""
        ins20_start_date = ""
        ins20_end_date = ""
        ins20_cost = ""
    if ins20 is not None:
        ins20_company = ins20.company
        ins20_telephone = ins20.telephone
        ins20_email = ins20.email 
        ins20_po_no = ins20.po_no
        ins20_reference_no = ins20.reference_no
        ins20_date_issued = ins20.date_issued
        ins20_start_date = ins20.start_date
        ins20_end_date = ins20.end_date
        ins20_cost = ins20.cost

    choices = ["M","S","F","L30","SUV","G15","G12","L3","SC","GR","DMC","GCM","CAC","CAI","D","G","A","M","R","NRC","NYR","NA","DNR"]

    real_value = ["Mitsubishi","Suzuki","Foton","L300 Exceed 2.5D MT","Super Carry UV","Gratour midi truck 1.5L",
        "Gratour midi truck 1.2L","L300 Exceed C/C","Suzuki CAB CHAS","Gratour midi","Diamond Motor Corporation",
        "Grand Canyon Multi Holdings, INC.","Cebu Autocentrale Corporation","Cherub Autodealer Inc.","Diesel",
        "Gas","Automatic","Manual","No Recieving Copy","Not Yet Release","Not Applicable","Did Not Recieve"]

    index = choices.index(data.brand) 
    data.brand = real_value[index]
    
    index = choices.index(data.make) 
    data.make = real_value[index]
    
    index = choices.index(data.series) 
    data.series = real_value[index]
    
    index = choices.index(data.dealer) 
    data.dealer = real_value[index]
    
    index = choices.index(data.fuel_type) 
    data.fuel_type = real_value[index]
    
    index = choices.index(data.transmission) 
    data.transmission = real_value[index]
    
    if data.plate_date == ("NRC" or "NYR" or "NA" or "DNR"): 
        index = choices.index(data.plate_date) 
        data.plate_date = real_value[index]
    else:
        pass
            
    if data.decals_date == ("NRC" or "NYR" or "NA" or "DNR"): 
        index = choices.index(data.decals_date) 
        data.decals_date = real_value[index]
    else:
        pass
            
    if data.tools_date == ("NRC" or "NYR" or "NA" or "DNR"): 
        index = choices.index(data.tools_date) 
        data.tools_date = real_value[index]
    else:
        pass
            
    if data.userManual_date == ("NRC" or "NYR" or "NA" or "DNR"): 
        index = choices.index(data.userManual_date) 
        data.userManual_date = real_value[index]
    else:
        pass
            
    if data.warrantyBook_date == ("NRC" or "NYR" or "NA" or "DNR"): 
        index = choices.index(data.warrantyBook_date) 
        data.warrantyBook_date = real_value[index]
    else:
        pass
            
    if data.unitKey_date == ("NRC" or "NYR" or "NA" or "DNR"): 
        index = choices.index(data.unitKey_date) 
        data.unitKey_date = real_value[index]
    else:
        pass
                
    if data.bodyKey_date == ("NRC" or "NYR" or "NA" or "DNR"): 
        index = choices.index(data.bodyKey_date) 
        data.bodyKey_date = real_value[index]
    else:
        pass
            
    if data.cigarettePlug_date == ("NRC" or "NYR" or "NA" or "DNR"): 
        index = choices.index(data.cigarettePlug_date) 
        data.cigarettePlug_date = real_value[index]
    else:
        pass
            
    if data.keychain_date == ("NRC" or "NYR" or "NA" or "DNR"): 
        index = choices.index(data.keychain_date) 
        data.keychain_date = real_value[index]
    else:
        pass
            
    if data.jack == ("NRC" or "NYR" or "NA" or "DNR"): 
        index = choices.index(data.jack) 
        data.jack = real_value[index]
    else:
        pass
            
    if data.wrench == ("NRC" or "NYR" or "NA" or "DNR"): 
        index = choices.index(data.wrench) 
        data.wrench = real_value[index]
    else:
        pass
            
    if data.fire_extinguisher == ("NRC" or "NYR" or "NA" or "DNR"): 
        index = choices.index(data.fire_extinguisher) 
        data.fire_extinguisher = real_value[index]
    else:
        pass
            
    if data.ewd_date == ("NRC" or "NYR" or "NA" or "DNR"): 
        index = choices.index(data.ewd_date) 
        data.ewd_date = real_value[index]
    else:
        pass
            
    if data.fan_date == ("NRC" or "NYR" or "NA" or "DNR"): 
        index = choices.index(data.fan_date) 
        data.fan_date = real_value[index]
    else:
        pass

    if data.operational == True:
        data.operational = "Yes"
    else:
        data.operational = "No"

    if data.status == "A":
        data.status = "Active"
    elif data.status == "M":
        data.status = "Maintenance"
    else:
        data.status = "Repair"
    

    # Define the data for each cell in the row 
    row = [
        data.car_id,
        data.vin_no,
        data.body_no,
        data.cs_no,
        data.plate_no,
        data.brand,
        data.release_year,
        data.make,
        data.series,
        data.body_type,
        data.color,
        data.dealer,
        data.dealer_phone,
        data.dealer_email,
        data.po_no,
        data.po_date,
        data.body_builder,
        data.fabricator,
        data.sale_price,
        data.vat_price,
        data.engine_no,
        data.battery_no,
        data.fuel_type,
        data.transmission,
        data.denomination,
        data.piston,
        data.cylinder,
        data.procuring_entity,
        data.capacity,
        data.gross_weight,
        data.net_weight,
        data.shipping_weight,
        data.net_capacity,
        data.lto_cr,
        data.cr_date,
        data.or_no,
        data.or_date,
        data.top_load,
        data.field_office,
        data.or_cr,
        data.permanent_loc,
        data.current_loc,
        data.vtf,
        data.permanent_status,
        data.delivery_location,
        data.deliver_date,
        data.si_no,
        data.dr_no,
        data.dr_codes,
        data.plate_date,
        data.decals_date,
        data.modified,
        data.ewd_date,
        data.tools_date,
        data.userManual_date,
        data.warrantyBook_date,
        data.unitKey_date,
        data.bodyKey_date,
        data.cigarettePlug_date,
        data.keychain_date,
        data.fan_date,
        data.jack,
        data.wrench,
        data.fire_extinguisher,

        contract_client_name,
        contract_contract_no,
        contract_start_date,
        contract_end_date,
        contract_bid_no,
        contract_bid_name,
        contract_bid_date,
        contract_cost,
        
        tpl_insurance_name,
        tpl_telephone,
        tpl_email,
        tpl_po_no,
        tpl_date_issued,
        tpl_start_date,
        tpl_end_date,
        tpl_cost,

        ins19_company,
        ins19_telephone,
        ins19_email,
        ins19_po_no,
        ins19_reference_no,
        ins19_date_issued,
        ins19_start_date,
        ins19_end_date,
        ins19_cost,
        
        ins20_company,
        ins20_telephone,
        ins20_email,
        ins20_po_no,
        ins20_reference_no,
        ins20_date_issued,
        ins20_start_date,
        ins20_end_date,
        ins20_cost,

        data.remarks,
        data.operational,
        data.status,
        data.date_updated,
        data.date_created,
    ]
    # Assign the data for each cell of the row 
    for col_num, cell_value in enumerate(row, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = cell_value
        cell.alignment = Alignment(horizontal='center')
        thin = Side(border_style="thin", color="000000")
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

workbook.save('.'+settings.MEDIA_URL+worksheet.title)
 

