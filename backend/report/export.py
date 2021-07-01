from datetime import datetime

from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .models import Cost


def export(inspection):
    datas = inspection
    workbook = Workbook()
    
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = '{date}-Inspections.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'))

    cell_1 = ["B1","H1","Q1","U1","AA1","AP1","AR1","BB1"]
    cell_2 = ["G1","P1","T1","Z1","AO1","AQ1","BA1","BF1"]
    header = ["VEHICLE INFOMATION","EXTERIOR","INTERIOR","ENGINE BAY","ELECTRICS",
            "WHEELS AND TIRES","GAS AND OIL","CHECKLIST REPORT"]
    color = ["00FFFF00","003366FF","00FFCC99","00C0C0C0","0000FF00","00FF0000","00800080","00FFFFFF"]

    for i in range(8):
        worksheet.merge_cells(cell_1[i] +":"+cell_2[i])

        cell = worksheet[cell_1[i]] 
        cell.value = header[i]
        
        cell.fill = PatternFill("solid", fgColor=color[i])
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(bold=True)
        thin = Side(border_style="thin", color="000000")
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # Define the titles for columns
    columns = [
        'ID',
        'Body No.',
        'Plate No.',
        'Make',
        'Current Location',
        "Mileage",
        "GPS",

        "Cleanliness",
        "Condition Rust",
        "Decals/Livery Intact",
        "Windows/Windscreen",
        "Rear Door",
        "Mirrors",
        "Roof Rack",
        "Rear Step",
        "Seats",
        "Seat Belts",
        "General Condition",
        "Vehicle Documents",
        
        "Engine Bay Cleanliness ",
        "Washer Fluids",
        "Coolant Level",
        "Brake Fluid Level",
        "Power Steering Fluid",
        "Liquid Leak Under Vehicle",

        "Main Beam",
        "Dipped Beam",
        "Side Lights",
        "Tail Lights",
        "Indicators",
        "Break Lights",
        "Reverse Lights",
        "Hazard Lights",
        "Rear Fog Light",
        "Interior Lights",
        "Screen Washer",
        "Wiper Blades",
        "Horn",
        "Radio/CD",
        "Front Fog Lights",

        "Gas Level",
        "Oil Level",

        "Tyres",
        "Front (Visual)",
        "Rear (Visual)",
        "Spare (Visual)",
        "Wheel Brace",
        "Jack",
        "Right Front",
        "Left Front",
        "Right Rear",
        "Left Rear",

        "Driver",
        "Edited By",
        "Notes",
        "Date Updated",
        "Date Created"
    ]
    row_num = 2

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.fill = PatternFill("solid", fgColor=color[i])
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(bold=True)
        thin = Side(border_style="thin", color="000000")
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        if col_num >= 2:
            cell.fill = PatternFill("solid", fgColor=color[0])
        if col_num >= 8:
            cell.fill = PatternFill("solid", fgColor=color[1])
        if col_num >= 17:
            cell.fill = PatternFill("solid", fgColor=color[2])
        if col_num >= 21:
            cell.fill = PatternFill("solid", fgColor=color[3])
        if col_num >= 27:
            cell.fill = PatternFill("solid", fgColor=color[4])
        if col_num >= 42:
            cell.fill = PatternFill("solid", fgColor=color[5])
        if col_num >= 44:
            cell.fill = PatternFill("solid", fgColor=color[6])
        if col_num >= 54:
            cell.fill = PatternFill("solid", fgColor=color[7])

    # Iterate through all movies
    for data in datas:
        row_num += 1
           
        choices = ["L30","SUV","G15","G12",True,False,None]
        real_value = ["L300 Exceed 2.5D MT","Super Carry UV","Gratour midi truck 1.5L",
            "Gratour midi truck 1.2L","Okay","Not Okay",None
            ]

        index = choices.index(data.body_no.make) 
        data.body_no.make = real_value[index]

        index = choices.index(data.cleanliness_exterior) 
        data.cleanliness_exterior = real_value[index]

        index = choices.index(data.condition_rust) 
        data.condition_rust = real_value[index]

        index = choices.index(data.decals) 
        data.decals = real_value[index]
            
        index = choices.index(data.windows) 
        data.windows = real_value[index]

        index = choices.index(data.rear_door) 
        data.rear_door = real_value[index]

        index = choices.index(data.mirror) 
        data.mirror = real_value[index]

        index = choices.index(data.roof_rack) 
        data.roof_rack = real_value[index]

        index = choices.index(data.rear_step) 
        data.rear_step = real_value[index]

        index = choices.index(data.seats) 
        data.seats = real_value[index]

        index = choices.index(data.seat_belts) 
        data.seat_belts = real_value[index]

        index = choices.index(data.general_condition) 
        data.general_condition = real_value[index]

        index = choices.index(data.vehicle_documents) 
        data.vehicle_documents = real_value[index]

        index = choices.index(data.main_beam) 
        data.main_beam = real_value[index]

        index = choices.index(data.dipped_beam) 
        data.dipped_beam = real_value[index]

        index = choices.index(data.side_lights) 
        data.side_lights = real_value[index]

        index = choices.index(data.tail_lights) 
        data.tail_lights = real_value[index]

        index = choices.index(data.indicators) 
        data.indicators = real_value[index]

        index = choices.index(data.break_lights) 
        data.break_lights = real_value[index]

        index = choices.index(data.reverse_lights) 
        data.reverse_lights = real_value[index]

        index = choices.index(data.hazard_light) 
        data.hazard_light = real_value[index]

        index = choices.index(data.rear_fog_lights) 
        data.rear_fog_lights = real_value[index]

        index = choices.index(data.interior_lights) 
        data.interior_lights = real_value[index]

        index = choices.index(data.screen_washer) 
        data.screen_washer = real_value[index]

        index = choices.index(data.wiper_blades) 
        data.wiper_blades = real_value[index]

        index = choices.index(data.horn) 
        data.horn = real_value[index]

        index = choices.index(data.radio) 
        data.radio = real_value[index]

        index = choices.index(data.front_fog_lights) 
        data.front_fog_lights = real_value[index]

        index = choices.index(data.liquid_leak) 
        data.liquid_leak = real_value[index]

        index = choices.index(data.cleanliness_engine_bay) 
        data.cleanliness_engine_bay = real_value[index]

        index = choices.index(data.washer_fluid) 
        data.washer_fluid = real_value[index]

        index = choices.index(data.coolant_level) 
        data.coolant_level = real_value[index]

        index = choices.index(data.brake_fluid_level) 
        data.brake_fluid_level = real_value[index]

        index = choices.index(data.power_steering_fluid) 
        data.power_steering_fluid = real_value[index]

        index = choices.index(data.tyres) 
        data.tyres = real_value[index]

        index = choices.index(data.front_visual) 
        data.front_visual = real_value[index]

        index = choices.index(data.rear_visual) 
        data.rear_visual = real_value[index]

        index = choices.index(data.spare_visual) 
        data.spare_visual = real_value[index]

        index = choices.index(data.wheel_brace) 
        data.wheel_brace = real_value[index]

        index = choices.index(data.jack) 
        data.jack = real_value[index]

        index = choices.index(data.front_right_wheel) 
        data.front_right_wheel = real_value[index]

        index = choices.index(data.front_left_wheel) 
        data.front_left_wheel = real_value[index]

        index = choices.index(data.rear_right_wheel) 
        data.rear_right_wheel = real_value[index]

        index = choices.index(data.rear_left_wheel) 
        data.rear_left_wheel = real_value[index]

        if data.edited_by != None:
            edited_by = data.edited_by.user_info.full_name
        else:
            edited_by = ""
        # Define the data for each cell in the row 
        row = [
            data.inspection_id,
            data.body_no.body_no,
            data.body_no.plate_no,
            data.body_no.make,
            data.body_no.current_loc,
            data.mileage,
            data.gps,

            data.cleanliness_exterior,
            data.condition_rust,
            data.decals,
            data.windows,
            data.rear_door,
            data.mirror,
            data.roof_rack,
            data.rear_step,
            data.seats,
            data.seat_belts,
            data.general_condition,
            data.vehicle_documents,
            
            data.cleanliness_engine_bay,
            data.washer_fluid,
            data.coolant_level,
            data.brake_fluid_level,
            data.power_steering_fluid,
            data.liquid_leak,

            data.main_beam,
            data.dipped_beam,
            data.side_lights,
            data.tail_lights,
            data.indicators,
            data.break_lights,
            data.reverse_lights,
            data.hazard_light,
            data.rear_fog_lights,
            data.interior_lights,
            data.screen_washer,
            data.wiper_blades,
            data.horn,
            data.radio,
            data.front_fog_lights,

            data.gas_level,
            data.oil_level,

            data.tyres,
            data.front_visual,
            data.rear_visual,
            data.spare_visual,
            data.wheel_brace,
            data.jack,
            data.front_right_wheel,
            data.front_left_wheel,
            data.rear_right_wheel,
            data.rear_left_wheel,

            data.driver.user_info.full_name,
            edited_by,
            data.notes,
            data.date_updated,
            data.date_created
        ]
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = Alignment(horizontal='center')
            thin = Side(border_style="thin", color="000000")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    try:
        workbook.save('.'+settings.MEDIA_URL+worksheet.title)
        return True
    except:
        return False

def repair_export(repair):
    datas = repair
    workbook = Workbook()
    
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = '{date}-Repair.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'))

    cell_1 = ["B1","D1","R1","X1","BE1","BK1"]
    cell_2 = ["C1","Q1","W1","BD1","BJ1","BL1"]
    header = ["JOB ORDER","VEHICLE INFOMATION","ACTUAL FINDINGS","COSTS","ACTION TAKEN","CHECKLIST REPORT"]
    color = ["00FFFF00","003366FF","00FFCC99","0000FF00","00C0C0C0","00FFFFFF"]

    for i in range(6):
        worksheet.merge_cells(cell_1[i] +":"+cell_2[i])
        cell = worksheet[cell_1[i]] 
        cell.value = header[i]
        cell.fill = PatternFill("solid", fgColor=color[i])
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(bold=True)
        thin = Side(border_style="thin", color="000000")
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # Define the titles for columns
    columns = [
        'ID',
        'Job Order No.',
        'Job Order Type',
        
        'Vin No.',
        'Body No.',
        'Plate No.',
        'Make',
        'Current Location',
        'Status',
        'Operational',
        'Schedule Date',

        'IR No.',
        'Incident Date',
        "Date Receive",
        "Site Poc",
        "Contact No.",

        "Incedent Details",
        "Diagnosed By",
        "Date Performed",
        "Generated By",
        "Noted By",
        "Actual Findings",
        "Remarks",

        "Parts",
        "Qty",
        "Cost",
        
        "Parts 2",
        "Qty 2",
        "Cost 2",
        
        "Parts 3",
        "Qty 3",
        "Cost 3",
        
        "Parts 4",
        "Qty 4",
        "Cost 4",
        
        "Parts 5",
        "Qty 5",
        "Cost 5",
        
        "Labor",
        "Qty L",
        "CostL ",
        
        "Labor 2",
        "Qty L2",
        "Cost L2",
        
        "Labor 3",
        "Qty L3",
        "Cost L3",
        
        "Labor 4",
        "Qty L4",
        "Cost L4",
        
        "Labor 5",
        "Qty L5",
        "Cost L5",

        "Total Parts Cost",
        "Total Labor Cost",
        "Total Estimate Cost",

        "Repair By",
        "Repair Date",
        "Date Done",
        "Status After Repair",
        "Action Taken",
        "Remarks",

        "Date Created",
        "Date Updated"
    ]
    row_num = 2

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.fill = PatternFill("solid", fgColor=color[i])
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(bold=True)
        thin = Side(border_style="thin", color="000000")
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        if col_num >= 2:
            cell.fill = PatternFill("solid", fgColor=color[0])
        if col_num >= 4:
            cell.fill = PatternFill("solid", fgColor=color[1])
        if col_num >= 18:
            cell.fill = PatternFill("solid", fgColor=color[2])
        if col_num >= 24:
            cell.fill = PatternFill("solid", fgColor=color[3])
        if col_num >= 57:
            cell.fill = PatternFill("solid", fgColor=color[4])
        if col_num >= 63:
            cell.fill = PatternFill("solid", fgColor=color[5])

    # Iterate through all movies
    for data in datas:
        row_num += 1
        costs = Cost.objects.filter(ro_no=data.repair_id)

        choices = ["L30","SUV","G15","G12",'A','M','R',True,False,None]
        real_value = ["L300 Exceed 2.5D MT","Super Carry UV","Gratour midi truck 1.5L",
            "Gratour midi truck 1.2L",'Active','Maintenance','Repair',"Repair","Inspection",None
            ]

        index = choices.index(data.job_order.type) 
        data.job_order.type = real_value[index]

        index = choices.index(data.job_order.task.body_no.make) 
        data.job_order.task.body_no.make = real_value[index]
        
        index = choices.index(data.job_order.task.body_no.status) 
        data.job_order.task.body_no.status = real_value[index]

        if data.job_order.task.body_no.operational == True:
            data.job_order.task.body_no.operational = "Operational"
        else:
            data.job_order.task.body_no.operational = "Non-Operational"

        if data.noted_by is not None:
            data.noted_by.username.user_info.full_name

        parts = ""
        qty = ""
        cost = ""
        parts2 = ""
        qty2 = ""
        cost2 = ""
        parts3 = ""
        qty3 = ""
        cost3 = ""
        parts4 = ""
        qty4 = ""
        cost4 = ""
        parts5 = ""
        qty5 = ""
        cost5 = ""

        labor = ""
        qtyl = ""
        costl = ""
        labor2 = ""
        qtyl2 = ""
        costl2 = ""
        labor3 = ""
        qtyl3 = ""
        costl3 = ""
        labor4 = ""
        qtyl4 = ""
        costl4 = ""
        labor5 = ""
        qtyl5 = ""
        costl5 = ""
        countP = 0
        countL = 0
        for i in range(len(costs)):
            if costs[i].cost_type == "P":
                countP+=1
                if countP == 1:
                    parts = costs[i].particulars
                    qty = costs[i].quantity
                    cost = costs[i].cost
                    print(costs[i].quantity)
                elif countP == 2:
                    parts2 = costs[i].particulars
                    qty2 = costs[i].quantity
                    cost2 = costs[i].cost
                    print(costs[i].quantity)
                elif countP == 3:
                    parts3 = costs[i].particulars
                    qty3 = costs[i].quantity
                    cost3 = costs[i].cost
                    print(costs[i].quantity)
                elif countP == 4:
                    parts4 = costs[i].particulars
                    qty4 = costs[i].quantity
                    cost4 = costs[i].cost
                elif countP == 5:
                    parts5 = costs[i].particulars
                    qty5 = costs[i].quantity
                    cost5 = costs[i].cost
            else:
                countL+=1
                if countL == 1:
                    labor = costs[i].particulars
                    qtyl = costs[i].quantity
                    costl = costs[i].cost
                elif countL == 2:
                    labor2 = costs[i].particulars
                    qtyl2 = costs[i].quantity
                    costl2 = costs[i].cost
                elif countL == 3:
                    labor3 = costs[i].particulars
                    qtyl3 = costs[i].quantity
                    costl3 = costs[i].cost
                elif countL == 4:
                    labor4 = costs[i].particulars
                    qtyl4 = costs[i].quantity
                    costl4 = costs[i].cost
                elif countL == 5:
                    labor5 = costs[i].particulars
                    qtyl5 = costs[i].quantity
                    costl5 = costs[i].cost

        print(data.total_estimate_cost)
        # Define the data for each cell in the row 
        row = [
            data.repair_id,
            data.job_order.job_no,
            data.job_order.type,

            data.job_order.task.body_no.vin_no,
            data.job_order.task.body_no.body_no,
            data.job_order.task.body_no.plate_no,
            data.job_order.task.body_no.make,
            data.job_order.task.body_no.current_loc,
            data.job_order.task.body_no.status,
            data.job_order.task.body_no.operational,
            data.job_order.task.schedule_date,

            data.ir_no,
            data.incident_date,
            data.date_receive,
            data.site_poc,
            data.contact_no,
            data.incident_details,
            
            str(data.diagnosed_by.user_info.full_name),
            data.perform_date,
            str(data.generated_by.user_info.full_name),
            str(data.noted_by),
            data.actual_findings,
            data.actual_remarks,
            
            parts,
            qty,
            cost,
            parts2,
            qty2,
            cost2,
            parts3,
            qty3,
            cost3,
            parts4,
            qty4,
            cost4,
            parts5,
            qty5,
            cost5,

            labor,
            qtyl,
            costl,
            labor2,
            qtyl2,
            costl2,
            labor3,
            qtyl3,
            costl3,
            labor4,
            qtyl4,
            costl4,
            labor5,
            qtyl5,
            costl5,

            data.total_parts_cost,
            data.total_labor_cost,
            data.total_parts_cost+data.total_labor_cost,
            
            str(data.repair_by.user_info.full_name),
            data.repair_date,
            data.date_done,
            data.status_repair,
            data.action_taken,
            data.remarks,

            data.date_updated,
            data.date_created,
            ]
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            c = cell.column
            if (c==26 or c==29 or c==32 or c==35 or c== 38 or c==56 or
                c==41 or c==44 or c==47 or c==50 or c==53 or c==54 or c==55 ):
                cell.number_format = "₱* #,##0.00"
            cell.value = cell_value
            cell.alignment = Alignment(horizontal='center')
            thin = Side(border_style="thin", color="000000")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    try:
        workbook.save('.'+settings.MEDIA_URL+worksheet.title)
        return True
    except:
        return False
